from openpyxl import load_workbook

def list_sheet_names(excel_path):
    """
    List all sheet names in an Excel file.
    """
    # Load the workbook
    wb = load_workbook(excel_path, read_only=True)
    # Get sheet names
    sheet_names = wb.sheetnames
    # Close the workbook
    wb.close()
    return sheet_names